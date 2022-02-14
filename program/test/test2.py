from openpyxl import Workbook
from openpyxl.styles import Font

wb=Workbook()
ws=wb.active
ws['A1']=10
cel=ws['A1']
cel.font=Font(size=20,bold=True,color='FF0000')
wb.save('../test2.xlsx')
